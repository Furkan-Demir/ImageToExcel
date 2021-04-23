from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill
from openpyxl.cell import cell
from PIL import Image 


def Resize(img):
    try:
        img = Image.open(img)

        new_img = img.resize((64,64), Image.ANTIALIAS)
        size = new_img.size
        if size[0] > size[1]:
            maxDimensionPixel = size[0]
        else:
            maxDimensionPixel = size[0]
        return maxDimensionPixel, new_img
    except:
        raise Exception("File not exist")

def excel(pixel,img):
    try:
        wb = Workbook()
        ws = wb.active
        for i in range(1,pixel-1):
            col_i = get_column_letter(i)

            ws.column_dimensions[col_i].width = 1
            ws.row_dimensions[i].height = 4


        for i in range(1,pixel-1):
            for k in range(1,pixel-1):
                r, g, b = img.getpixel((i,k))
                hexColor = "{0:02X}{1:02X}{2:02X}".format(r, g, b)
                myCell = ws.cell(row=k,column=i)
                ws[myCell.coordinate].fill = PatternFill(fgColor=hexColor, fill_type = "solid")

        wb.save("output.xlsx")
        return True
    except Exception as e:
        print(e)
        return False



if __name__ == '__main__':
    print("Welcome to Image To Excel Python Program\nType a file name (E.g input.jpeg ) \nIt's have to exist same this python's script directory")

    img = input("File name: ")
    pixel, img = Resize(img)
    status = excel(pixel,img)
    if status == True:
        print("Success!")
    else:
        print("Fail :(")


